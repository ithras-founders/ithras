"""
Analytics export - PDF and Excel generation.
"""
import io
import json
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Optional, List, Any
from pydantic import BaseModel

from app.modules.shared import database

router = APIRouter(prefix="/api/v1/analytics/export", tags=["analytics-export"])


class ExportRequest(BaseModel):
    format: str  # "pdf" | "xlsx"
    report_id: Optional[int] = None
    query: Optional[str] = None
    params: Optional[List[Any]] = None
    chart_config: Optional[dict] = None
    columns: Optional[List[str]] = None
    rows: Optional[List[List[Any]]] = None


@router.post("")
def export_data(req: ExportRequest, db: Session = Depends(database.get_db)):
    """Export report/data as PDF or Excel."""
    if req.format not in ("pdf", "xlsx"):
        raise HTTPException(status_code=400, detail="format must be 'pdf' or 'xlsx'")

    columns = req.columns or []
    rows = req.rows or []

    if req.report_id and (not columns or not rows):
        # Load report and execute query
        report_row = db.execute(text(
            "SELECT query, params_json FROM analytics_reports WHERE id = :id"
        ), {"id": req.report_id}).fetchone()
        if not report_row:
            raise HTTPException(status_code=404, detail="Report not found")
        query, params_json = report_row[0], report_row[1]
        params = json.loads(params_json) if params_json else []
        try:
            result = db.execute(text(query), params or [])
            if result.returns_rows:
                rows_data = result.fetchmany(10000)
                columns = list(result.keys())
                rows = [[str(v) if hasattr(v, 'isoformat') else v for v in r] for r in rows_data]
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    if req.query and (not columns or not rows):
        try:
            result = db.execute(text(req.query), req.params or [])
            if result.returns_rows:
                rows_data = result.fetchmany(10000)
                columns = list(result.keys())
                rows = [[str(v) if hasattr(v, 'isoformat') else v for v in r] for r in rows_data]
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    if req.format == "xlsx":
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed for Excel export")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        for c, col in enumerate(columns, 1):
            ws.cell(row=1, column=c, value=col)
        for r, row_data in enumerate(rows, 2):
            for c, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c, value=val)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"analytics_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

    if req.format == "pdf":
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            raise HTTPException(status_code=500, detail="reportlab not installed for PDF export")
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter)
        elements = []
        if columns:
            data = [columns] + rows
            t = Table(data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(t)
        doc.build(elements)
        buf.seek(0)
        filename = f"analytics_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})

    raise HTTPException(status_code=400, detail="No data to export")
